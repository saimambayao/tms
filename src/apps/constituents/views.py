from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, UpdateView, CreateView, TemplateView
from django.views import View # Import the View class
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Avg, F, Sum, Case, When, Value, CharField
from django.db.models.functions import Concat
import datetime
import logging
import csv
import json
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from .models import Constituent, ConstituentInteraction, ConstituentGroup
from .member_models import BMParliamentMember
from .forms import ExcelUploadForm
from apps.users.models import User

logger = logging.getLogger(__name__)

class StaffRequiredMixin(UserPassesTestMixin):
    """
    Mixin that checks if the user is a staff member or above.
    """
    def test_func(self):
        # Superusers should always have access
        if self.request.user.is_superuser:
            return True
        return self.request.user.is_authenticated and self.request.user.is_staff_or_above()


class DatabaseAccessMixin(LoginRequiredMixin, UserPassesTestMixin):
    """
    Mixin that checks if the user has access to the Database of Registrants.
    Authorized roles: Superuser, MP, Chief of Staff, System Admin, and Coordinator
    """
    
    def test_func(self):
        authorized_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        return (self.request.user.is_superuser or 
                self.request.user.user_type in authorized_roles)
    
    def handle_no_permission(self):
        """Handle permission denied cases appropriately."""
        if not self.request.user.is_authenticated:
            # Unauthenticated users should be redirected to login
            return super().handle_no_permission()
        else:
            # Authenticated but unauthorized users go to home page
            messages.warning(self.request, 'You do not have permission to access this page.')
            return redirect('home')


# Database of Registrants Views
class DatabaseRegistrantsView(DatabaseAccessMixin, ListView):
    """
    Comprehensive Database of Registrants for authorized users.
    Shows all BMParliamentMember registrations with search, filtering, and editing capabilities.
    """
    model = BMParliamentMember
    template_name = 'constituents/database_registrants.html'
    context_object_name = 'registrants'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = BMParliamentMember.objects.all().select_related('user', 'approved_by')
        
        # Search functionality
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(middle_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(contact_number__icontains=search_query) |
                Q(address_municipality__icontains=search_query) |
                Q(address_province__icontains=search_query) |
                Q(voter_address_municipality__icontains=search_query) |
                Q(voter_address_province__icontains=search_query)
            )
        
        # Province filter
        province_filter = self.request.GET.get('province')
        if province_filter:
            queryset = queryset.filter(address_province=province_filter)
        
        # Municipality filter
        municipality_filter = self.request.GET.get('municipality')
        if municipality_filter:
            queryset = queryset.filter(address_municipality=municipality_filter)
        
        # Approval status filter
        approval_status = self.request.GET.get('approval_status')
        if approval_status:
            queryset = queryset.filter(status=approval_status)
        
        # Sector filter
        sector_filter = self.request.GET.get('sector')
        if sector_filter:
            queryset = queryset.filter(sector=sector_filter)
        
        # Age range filter
        age_filter = self.request.GET.get('age_range')
        if age_filter:
            if age_filter == '18-25':
                queryset = queryset.filter(age__range=(18, 25))
            elif age_filter == '26-35':
                queryset = queryset.filter(age__range=(26, 35))
            elif age_filter == '36-50':
                queryset = queryset.filter(age__range=(36, 50))
            elif age_filter == '51-65':
                queryset = queryset.filter(age__range=(51, 65))
            elif age_filter == '65+':
                queryset = queryset.filter(age__gte=65)
        
        # Education filter
        education_filter = self.request.GET.get('education')
        if education_filter:
            queryset = queryset.filter(highest_education=education_filter)
        
        # Gender filter
        gender_filter = self.request.GET.get('gender')
        if gender_filter:
            queryset = queryset.filter(sex=gender_filter)
        
        # Date range filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date_of_application__gte=date_from)
        if date_to:
            queryset = queryset.filter(date_of_application__lte=date_to)
        
        # Sorting
        sort_by = self.request.GET.get('sort_by', '-date_of_application')
        valid_sort_fields = [
            'first_name', '-first_name', 'last_name', '-last_name',
            'date_of_application', '-date_of_application', 'age', '-age',
            'address_province', '-address_province', 'address_municipality', '-address_municipality',
            'status', '-status', 'sector', '-sector'
        ]
        
        if sort_by in valid_sort_fields:
            if sort_by in ['first_name', '-first_name', 'last_name', '-last_name']:
                # Handle name sorting properly
                if sort_by == 'first_name':
                    queryset = queryset.order_by('first_name', 'last_name')
                elif sort_by == '-first_name':
                    queryset = queryset.order_by('-first_name', '-last_name')
                elif sort_by == 'last_name':
                    queryset = queryset.order_by('last_name', 'first_name')
                elif sort_by == '-last_name':
                    queryset = queryset.order_by('-last_name', '-first_name')
            else:
                queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('-date_of_application')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics for dashboard cards
        context['total_registrants'] = BMParliamentMember.objects.count()
        context['approved_count'] = BMParliamentMember.objects.filter(status='approved').count()
        context['pending_count'] = BMParliamentMember.objects.filter(status='pending').count()
        context['incomplete_count'] = BMParliamentMember.objects.filter(status='incomplete').count()
        context['non_compliant_count'] = BMParliamentMember.objects.filter(status='non_compliant').count()
        context['denied_count'] = BMParliamentMember.objects.filter(status='denied').count() # Keep for backward compatibility if needed, or remove if 'denied' is fully replaced
        context['this_month_count'] = BMParliamentMember.objects.filter(
            date_of_application__gte=timezone.now().replace(day=1, hour=0, minute=0)
        ).count()
        
        # Province/Municipality choices for filters - using same list as register page
        context['provinces'] = ['Maguindanao del Sur']
        
        context['municipalities'] = [
            'Ampatuan',
            'Buluan',
            'Datu Abdullah Sangki',
            'Datu Anggal Midtimbang',
            'Datu Blah T. Sinsuat',
            'Datu Hoffer Ampatuan',
            'Datu Odin Sinsuat',
            'Datu Paglas',
            'Datu Piang',
            'Datu Salibo',
            'Datu Saudi-Ampatuan',
            'Datu Unsay',
            'General Salipada K. Pendatun',
            'Guindulungan',
            'Mamasapano',
            'Mangudadatu',
            'Pagalungan',
            'Paglat',
            'Pandag',
            'Rajah Buayan',
            'Shariff Aguak',
            'Shariff Saydona Mustapha',
            'South Upi',
            'Sultan Kudarat',
            'Sultan Mastura',
            'Sultan sa Barongis',
            'Talayan',
            'Talitay'
        ]
        
        # Sector choices
        context['sectors'] = BMParliamentMember.SECTOR_CHOICES
        
        # Education choices
        context['education_levels'] = BMParliamentMember.EDUCATION_CHOICES
        
        # Age range options
        context['age_ranges'] = [
            ('18-25', '18-25 years'),
            ('26-35', '26-35 years'),
            ('36-50', '36-50 years'),
            ('51-65', '51-65 years'),
            ('65+', '65+ years'),
        ]
        
        # Gender options
        context['genders'] = BMParliamentMember.SEX_CHOICES
        
        # Current filter values for maintaining state
        context['current_filters'] = {
            'search': self.request.GET.get('search', ''),
            'province': self.request.GET.get('province', ''),
            'municipality': self.request.GET.get('municipality', ''),
            'approval_status': self.request.GET.get('approval_status', ''),
            'sector': self.request.GET.get('sector', ''),
            'age_range': self.request.GET.get('age_range', ''),
            'education': self.request.GET.get('education', ''),
            'gender': self.request.GET.get('gender', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'sort_by': self.request.GET.get('sort_by', '-date_of_application'),
        }
        
        # Filter parameters for pagination
        filter_params = self.request.GET.copy()
        if 'page' in filter_params:
            del filter_params['page']
        context['filter_params'] = filter_params
        
        # User role for template conditional rendering
        context['user_role'] = self.request.user.user_type
        # Include coordinator in approval permissions as per user requirements
        context['can_approve'] = (self.request.user.is_superuser or 
                                 self.request.user.user_type in ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator'])
        
        # Location-based analytics
        location_stats = self.get_queryset().values('address_province').annotate(
            total=Count('id'),
            approved_count=Count(Case(When(status='approved', then=1))),
            pending_count=Count(Case(When(status='pending', then=1))),
            incomplete_count=Count(Case(When(status='incomplete', then=1)))
        ).order_by('address_province')

        # Get municipality counts separately
        municipality_stats = self.get_queryset().values(
            'address_province', 'address_municipality'
        ).annotate(
            count=Count('id')
        ).order_by('address_province', 'address_municipality')

        province_stats = {}
        for stat in location_stats:
            province = stat['address_province']
            province_stats[province] = {
                'total': stat['total'],
                'approved_count': stat['approved_count'],
                'pending_count': stat['pending_count'],
                'incomplete_count': stat['incomplete_count'],
                'municipalities': {}
            }
        
        for stat in municipality_stats:
            province = stat['address_province']
            municipality = stat['address_municipality']
            count = stat['count']
            if province in province_stats:
                province_stats[province]['municipalities'][municipality] = count
            
        context['province_stats'] = province_stats
        
        return context


class ExcelNameCheckView(DatabaseAccessMixin, View):
    """
    View to upload Excel file with names and check against database for matches.
    Generates Excel output with results showing matches and findings.
    """
    template_name = 'constituents/excel_name_check.html'

    def get(self, request, *args, **kwargs):
        """Display the Excel upload form."""
        form = ExcelUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        """Process the uploaded Excel file and generate results."""
        form = ExcelUploadForm(request.POST, request.FILES)

        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        try:
            # Import pandas locally to avoid module-level import issues
            import pandas as pd

            # Get the uploaded DataFrame
            df = form.cleaned_data['dataframe']

            # Process the names and check for matches
            results = self.process_names_and_check_matches(df)

            # Generate Excel output
            excel_response = self.generate_excel_output(results)

            return excel_response

        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            messages.error(request, f"Error processing file: {str(e)}")
            return render(request, self.template_name, {'form': form})

    def process_names_and_check_matches(self, df):
        """
        Process the DataFrame containing names and check for matches in the database.
        Returns a list of results with match information.
        """
        # Import pandas locally to avoid module-level import issues
        import pandas as pd

        results = []

        # Check if we have separate name columns that should be combined
        first_name_col = None
        middle_name_col = None
        last_name_col = None
        full_name_col = None

        # Look for separate name columns
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'first' in col_lower and 'name' in col_lower:
                first_name_col = col
            elif 'middle' in col_lower and 'name' in col_lower:
                middle_name_col = col
            elif 'last' in col_lower and 'name' in col_lower:
                last_name_col = col
            elif any(name_type in col_lower for name_type in ['name', 'full_name', 'fullname', 'person_name', 'names']):
                full_name_col = col

        # Process each row
        for idx, row in df.iterrows():
            # Determine the original name to display
            if full_name_col:
                # Use the full name column if available
                original_name = str(row[full_name_col]).strip()
            elif first_name_col or middle_name_col or last_name_col:
                # Combine separate name columns
                name_parts = []
                if first_name_col and pd.notna(row[first_name_col]):
                    name_parts.append(str(row[first_name_col]).strip())
                if middle_name_col and pd.notna(row[middle_name_col]):
                    name_parts.append(str(row[middle_name_col]).strip())
                if last_name_col and pd.notna(row[last_name_col]):
                    name_parts.append(str(row[last_name_col]).strip())

                original_name = ' '.join(name_parts)
            else:
                # Fallback to first column
                original_name = str(row[df.columns[0]]).strip()

            if not original_name or original_name.lower() in ['nan', 'none', '']:
                continue

            # Search for matches in the database using the original name
            matches = self.find_name_matches(original_name)

            # Prepare result entry - keep the original name exactly as it appears in Excel
            result = {
                'original_name': original_name,  # Full name from Excel file (combined if separate columns)
                'matches_found': len(matches) > 0,
                'match_count': len(matches),
                'matches': matches,
                'other_databases': self.check_other_databases(original_name)
            }

            results.append(result)

        # Sort results alphabetically by original name
        results.sort(key=lambda x: x['original_name'].lower())

        return results

    def find_name_matches(self, name):
        """
        Search for precise name matches in both BMParliamentMember and Constituent databases.
        Parses full names properly and uses exact matching for accuracy.
        Returns a list of matching member/constituent details from unified databases.
        """
        matches = []

        # Normalize the search name - remove extra spaces and handle case
        search_name = ' '.join(name.split()).strip()

        # Parse the name into components - handle both "First Middle Last" and "Last, First Middle" formats
        parsed_names = self.parse_full_name(search_name)

        if not parsed_names:
            return matches

        first_name = parsed_names.get('first_name', '').strip()
        middle_name = parsed_names.get('middle_name', '').strip()
        last_name = parsed_names.get('last_name', '').strip()

        logger.info(f"Parsed name - First: '{first_name}', Middle: '{middle_name}', Last: '{last_name}'")

        # Build precise search queries - only exact matches, no partial fallbacks
        search_queries = []

        # Exact full name match (most precise) - require both first and last names
        if first_name and last_name:
            # Primary exact match: first + last name
            search_queries.append(Q(first_name__iexact=first_name, last_name__iexact=last_name))

            # If middle name provided, also check exact match with middle name
            if middle_name:
                search_queries.append(Q(first_name__iexact=first_name, middle_name__iexact=middle_name, last_name__iexact=last_name))

        # Combine all search conditions
        if search_queries:
            combined_query = search_queries[0]
            for query in search_queries[1:]:
                combined_query |= query

            # Search BMParliamentMember database
            member_queryset = BMParliamentMember.objects.filter(combined_query).distinct()

            for member in member_queryset[:10]:  # Allow more matches since we're more precise
                # Additional validation - check if names actually match closely
                if self.names_match_closely(member.first_name, member.middle_name, member.last_name, first_name, middle_name, last_name):
                    match_info = {
                        'full_name': member.get_full_name(),
                        'member_id': member.member_id or 'N/A',
                        'status': member.get_status_display(),
                        'sector': member.get_sector_display(),
                        'province': member.address_province,
                        'municipality': member.address_municipality,
                        'contact_number': member.contact_number,
                        'email': member.email,
                        'date_of_application': member.date_of_application.strftime('%Y-%m-%d') if member.date_of_application else '',
                        'database': 'BM Parliament Members',
                        'database_type': 'member'
                    }
                    matches.append(match_info)

        # Search Constituent database with same logic - only exact matches
        constituent_search_queries = []

        if first_name and last_name:
            # Primary exact match: first + last name
            constituent_search_queries.append(Q(user__first_name__iexact=first_name, user__last_name__iexact=last_name))

        if constituent_search_queries:
            constituent_combined_query = constituent_search_queries[0]
            for query in constituent_search_queries[1:]:
                constituent_combined_query |= query

            member_queryset = Constituent.objects.select_related('user').filter(constituent_combined_query).distinct()

            for constituent in member_queryset[:10]:
                # Additional validation for constituents
                if self.names_match_closely(constituent.user.first_name, '', constituent.user.last_name, first_name, middle_name, last_name):
                    match_info = {
                        'full_name': constituent.user.get_full_name() or constituent.user.username,
                        'member_id': 'N/A',
                        'status': 'Constituent',
                        'sector': 'General',
                        'province': 'N/A',
                        'municipality': 'N/A',
                        'contact_number': constituent.alternate_contact or 'N/A',
                        'email': constituent.user.email,
                        'date_of_application': constituent.created_at.strftime('%Y-%m-%d') if constituent.created_at else '',
                        'database': 'Constituents Database',
                        'database_type': 'constituent'
                    }
                    matches.append(match_info)

        return matches

    def parse_full_name(self, full_name):
        """
        Parse a full name string into first, middle, and last name components.
        Handles various formats: "First Middle Last", "Last, First Middle", etc.
        """
        if not full_name or not full_name.strip():
            return None

        name = full_name.strip()

        # Check if it's in "Last, First Middle" format
        if ',' in name:
            parts = name.split(',', 1)
            if len(parts) == 2:
                last_name = parts[0].strip()
                first_middle = parts[1].strip().split()
                first_name = first_middle[0] if first_middle else ''
                middle_name = ' '.join(first_middle[1:]) if len(first_middle) > 1 else ''
            else:
                # Fallback parsing
                name_parts = name.replace(',', '').split()
                first_name = name_parts[0] if name_parts else ''
                middle_name = ' '.join(name_parts[1:-1]) if len(name_parts) > 2 else ''
                last_name = name_parts[-1] if len(name_parts) > 1 else ''
        else:
            # Assume "First Middle Last" format
            name_parts = name.split()
            if len(name_parts) >= 2:
                first_name = name_parts[0]
                last_name = name_parts[-1]
                middle_name = ' '.join(name_parts[1:-1]) if len(name_parts) > 2 else ''
            else:
                # Single name - treat as first name
                first_name = name
                middle_name = ''
                last_name = ''

        return {
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name
        }

    def names_match_closely(self, db_first, db_middle, db_last, search_first, search_middle, search_last):
        """
        Check if database names match search names closely enough.
        Uses case-insensitive comparison and allows for minor variations.
        """
        # Normalize all names for comparison
        db_first = (db_first or '').strip().lower()
        db_middle = (db_middle or '').strip().lower()
        db_last = (db_last or '').strip().lower()
        search_first = (search_first or '').strip().lower()
        search_middle = (search_middle or '').strip().lower()
        search_last = (search_last or '').strip().lower()

        # Must match last name exactly (most important identifier)
        if search_last and db_last != search_last:
            return False

        # Must match first name exactly
        if search_first and db_first != search_first:
            return False

        # Middle name should match if provided in search
        if search_middle and db_middle != search_middle:
            return False

        return True

    def check_other_databases(self, name):
        """
        Check for presence in other databases/systems.
        This is a placeholder for future integration with other systems.
        """
        # For now, return a simple structure
        # In a real implementation, this would check other databases/APIs
        other_databases = {
            'notion_database': False,  # Placeholder
            'external_system_a': False,  # Placeholder
            'external_system_b': False,  # Placeholder
        }

        # Simple logic: if name contains certain patterns, mark as found
        # This is just for demonstration
        search_name = name.lower()
        if 'test' in search_name or 'sample' in search_name:
            other_databases['notion_database'] = True

        return other_databases

    def generate_excel_output(self, results):
        """
        Generate Excel file with the results.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Name Check Results"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        match_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        no_match_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        # Write headers
        headers = [
            'Original Name', 'Matches Found', 'Match Count', 'Database Matches',
            'Database Source', 'Notion Database', 'External System A', 'External System B'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data
        row_num = 2
        for result in results:
            # Basic info
            ws.cell(row=row_num, column=1, value=result['original_name'])
            ws.cell(row=row_num, column=2, value='Yes' if result['matches_found'] else 'No')
            ws.cell(row=row_num, column=3, value=result['match_count'])

            # Database matches (concatenated)
            if result['matches']:
                match_details = []
                for match in result['matches'][:3]:  # Limit to first 3 matches
                    match_str = f"{match['full_name']} ({match['member_id']}, {match['status']})"
                    match_details.append(match_str)
                ws.cell(row=row_num, column=4, value=' | '.join(match_details))
            else:
                ws.cell(row=row_num, column=4, value='No matches found')

            # Database source
            if result['matches']:
                database_sources = list(set(match['database'] for match in result['matches']))
                ws.cell(row=row_num, column=5, value=' | '.join(database_sources))
            else:
                ws.cell(row=row_num, column=5, value='No matches found')

            # Other databases
            other_db = result['other_databases']
            ws.cell(row=row_num, column=6, value='Yes' if other_db.get('notion_database') else 'No')
            ws.cell(row=row_num, column=7, value='Yes' if other_db.get('external_system_a') else 'No')
            ws.cell(row=row_num, column=8, value='Yes' if other_db.get('external_system_b') else 'No')

            # Apply row styling
            fill = match_fill if result['matches_found'] else no_match_fill
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = fill

            row_num += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="name_check_results_{timestamp}.xlsx"'

        return response


class ExportRegistrantsCSVView(DatabaseAccessMixin, View):
    """
    View to export filtered BMParliamentMember data as a CSV file.
    """
    def get(self, request, *args, **kwargs):
        queryset = BMParliamentMember.objects.all().select_related('user')

        # Apply filters similar to DatabaseRegistrantsView
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(middle_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(contact_number__icontains=search_query) |
                Q(address_municipality__icontains=search_query) |
                Q(address_province__icontains=search_query) |
                Q(voter_address_municipality__icontains=search_query) |
                Q(voter_address_province__icontains=search_query)
            )

        province_filter = self.request.GET.get('province')
        if province_filter:
            queryset = queryset.filter(address_province=province_filter)

        municipality_filter = self.request.GET.get('municipality')
        if municipality_filter:
            queryset = queryset.filter(address_municipality=municipality_filter)

        approval_status = self.request.GET.get('approval_status')
        if approval_status:
            queryset = queryset.filter(status=approval_status)

        sector_filter = self.request.GET.get('sector')
        if sector_filter:
            queryset = queryset.filter(sector=sector_filter)

        age_filter = self.request.GET.get('age_range')
        if age_filter:
            if age_filter == '18-25':
                queryset = queryset.filter(age__range=(18, 25))
            elif age_filter == '26-35':
                queryset = queryset.filter(age__range=(26, 35))
            elif age_filter == '36-50':
                queryset = queryset.filter(age__range=(36, 50))
            elif age_filter == '51-65':
                queryset = queryset.filter(age__range=(51, 65))
            elif age_filter == '65+':
                queryset = queryset.filter(age__gte=65)

        education_filter = self.request.GET.get('education')
        if education_filter:
            queryset = queryset.filter(highest_education=education_filter)

        gender_filter = self.request.GET.get('gender')
        if gender_filter:
            queryset = queryset.filter(sex=gender_filter)

        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date_of_application__gte=date_from)
        if date_to:
            queryset = queryset.filter(date_of_application__lte=date_to)

        # Sort alphabetically by last name, then first name
        queryset = queryset.order_by('last_name', 'first_name')

        # Prepare CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registrants_export.csv"'

        writer = csv.writer(response)
        writer.writerow(['Full Name', 'Username', 'Phone Number', 'Email', 'Status']) # CSV Header

        for member in queryset:
            # Format name as "Last Name, First Name Middle Name"
            if member.middle_name:
                full_name = f"{member.last_name}, {member.first_name} {member.middle_name}"
            else:
                full_name = f"{member.last_name}, {member.first_name}"

            writer.writerow([
                full_name,
                member.user.username if member.user else '',
                member.contact_number,
                member.email,
                member.get_status_display()
            ])
        return response


class DatabaseRegistrantDetailView(DatabaseAccessMixin, DetailView):
    """
    Detailed view of a single registrant with editing capabilities.
    """
    model = BMParliamentMember
    template_name = 'constituents/database_registrant_detail.html'
    context_object_name = 'registrant'

    def get_queryset(self):
        return super().get_queryset().select_related('user', 'approved_by')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ensure the registrant object is reloaded from the database to get the latest status
        self.object.refresh_from_db()
        context['registrant'] = self.object # Update context with the refreshed object
        
        logger.info(f"Registrant {self.object.pk} status in DatabaseRegistrantDetailView: {self.object.status}")
        
        # Include coordinator in edit and approval permissions as per user requirements
        authorized_edit_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        context['can_edit'] = (self.request.user.is_superuser or 
                              self.request.user.user_type in authorized_edit_roles)
        context['can_approve'] = (self.request.user.is_superuser or 
                                 self.request.user.user_type in authorized_edit_roles)
        context['user_role'] = self.request.user.user_type
        return context


class RegistrantApproveView(DatabaseAccessMixin, View):
    """
    View to approve a registrant directly.
    """
    def post(self, request, pk, *args, **kwargs):
        registrant = get_object_or_404(BMParliamentMember, pk=pk)
        
        # Check permissions
        authorized_approve_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        if not (request.user.is_superuser or request.user.user_type in authorized_approve_roles):
            messages.error(request, 'You do not have permission to approve registrants.')
            return redirect('database_registrant_detail', pk=pk)
        
        if registrant.status != 'approved':
            registrant.status = 'approved'
            registrant.approved_by = request.user
            registrant.approved_date = timezone.now().date()
            registrant.save()
            messages.success(request, f"Registrant {registrant.get_full_name()} has been approved.")
        else:
            messages.info(request, f"Registrant {registrant.get_full_name()} is already approved.")
        
        logger.info(f"Registrant {registrant.pk} status after approval attempt: {registrant.status}")
        return redirect('database_registrant_detail', pk=pk)


class RegistrantMarkIncompleteView(DatabaseAccessMixin, View):
    """
    View to mark a registrant as incomplete.
    """
    def post(self, request, pk, *args, **kwargs):
        registrant = get_object_or_404(BMParliamentMember, pk=pk)

        # Check permissions - same as approve roles for now
        authorized_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        if not (request.user.is_superuser or request.user.user_type in authorized_roles):
            messages.error(request, 'You do not have permission to mark registrants as incomplete.')
            return redirect('database_registrant_detail', pk=pk)

        if registrant.status != 'incomplete':
            registrant.status = 'incomplete'
            registrant.approved_by = None # Clear approved_by if marking incomplete
            registrant.approved_date = None # Clear approved_date if marking incomplete
            registrant.save()
            messages.success(request, f"Registrant {registrant.get_full_name()} has been marked as incomplete.")
        else:
            messages.info(request, f"Registrant {registrant.get_full_name()} is already marked as incomplete.")

        logger.info(f"Registrant {registrant.pk} status after incomplete attempt: {registrant.status}")
        return redirect('database_registrant_detail', pk=pk)


class RegistrantMarkNonCompliantView(DatabaseAccessMixin, View):
    """
    View to mark a registrant as non-compliant.
    """
    def post(self, request, pk, *args, **kwargs):
        registrant = get_object_or_404(BMParliamentMember, pk=pk)

        # Check permissions - same as approve roles for now
        authorized_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        if not (request.user.is_superuser or request.user.user_type in authorized_roles):
            messages.error(request, 'You do not have permission to mark registrants as non-compliant.')
            return redirect('database_registrant_detail', pk=pk)

        if registrant.status != 'non_compliant':
            registrant.status = 'non_compliant'
            registrant.approved_by = None # Clear approved_by if marking non-compliant
            registrant.approved_date = None # Clear approved_date if marking non-compliant
            registrant.save()
            messages.success(request, f"Registrant {registrant.get_full_name()} has been marked as non-compliant.")
        else:
            messages.info(request, f"Registrant {registrant.get_full_name()} is already marked as non-compliant.")

        logger.info(f"Registrant {registrant.pk} status after non-compliant attempt: {registrant.status}")
        return redirect('database_registrant_detail', pk=pk)


class DatabaseRegistrantUpdateView(DatabaseAccessMixin, UpdateView):
    """
    Update view for registrant information.
    """
    model = BMParliamentMember
    template_name = 'constituents/database_registrant_edit.html'
    fields = [
        'first_name', 'last_name', 'middle_name', 'email', 'contact_number', 'age', 'sex',
        'address_barangay', 'address_municipality', 'address_province',
        'voter_address_barangay', 'voter_address_municipality', 'voter_address_province',
        'sector', 'highest_education', 'school_graduated', 'eligibility',
        'voter_id_photo', # Added for verification document update
        'status'
    ]
    
    def test_func(self):
        # Check parent access first (DatabaseAccessMixin)
        if not super().test_func():
            return False
        
        # Check edit-specific access - include coordinator as per user requirements
        edit_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        return (self.request.user.is_superuser or 
                self.request.user.user_type in edit_roles)
    
    def form_valid(self, form):
        # Set approval information if status is being changed to approved
        if form.cleaned_data.get('status') == 'approved' and self.object.status != 'approved':
            form.instance.approved_by = self.request.user
            form.instance.approved_date = timezone.now().date()
        
        messages.success(
            self.request, 
            f"Registrant {self.object.get_full_name()} updated successfully."
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('database_registrant_detail', kwargs={'pk': self.object.pk})


class DatabaseRegistrantDeleteView(DatabaseAccessMixin, UserPassesTestMixin, DetailView):
    """
    View to delete a registrant.
    """
    model = BMParliamentMember
    template_name = 'constituents/database_registrant_confirm_delete.html'
    context_object_name = 'registrant'

    def test_func(self):
        # Check parent access first (DatabaseAccessMixin)
        if not super().test_func():
            return False
        
        # Only superusers, admins, and chief of staff can delete
        delete_roles = ['superuser', 'chief_of_staff', 'admin']
        return (self.request.user.is_superuser or 
                self.request.user.user_type in delete_roles)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        registrant_name = self.object.get_full_name()
        self.object.delete()
        messages.success(self.request, f"Registrant {registrant_name} deleted successfully.")
        return redirect('database_registrants')


class AdminPasswordResetView(DatabaseAccessMixin, View):
    """
    Allows an admin to reset a user's password to a default value.
    """
    def get(self, request, user_pk):
        user_to_reset = get_object_or_404(User, pk=user_pk)
        
        try:
            user_to_reset.set_password('bmparliament123')
            user_to_reset.save()
            messages.success(request, f"Password for {user_to_reset.username} has been reset to the default.")
        except Exception as e:
            messages.error(request, f"Could not reset password: {e}")

        # Redirect back to the registrant detail page
        member = get_object_or_404(BMParliamentMember, user=user_to_reset)
        return redirect('database_registrant_detail', pk=member.pk)


class AddRegistrantGroupView(DatabaseAccessMixin, TemplateView):
    """
    View to add a list of existing members to a group with remarks.
    """
    template_name = 'constituents/add_registrant_group.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_members'] = BMParliamentMember.objects.all().order_by('last_name', 'first_name')
        
        # Pass filter choices to the template for dynamic filtering
        context['sectors'] = BMParliamentMember.SECTOR_CHOICES
        context['education_levels'] = BMParliamentMember.EDUCATION_CHOICES
        context['age_ranges'] = [
            ('18-25', '18-25 years'),
            ('26-35', '26-35 years'),
            ('36-50', '36-50 years'),
            ('51-65', '51-65 years'),
            ('65+', '65+ years'),
        ]
        context['genders'] = BMParliamentMember.SEX_CHOICES
        
        return context

    def post(self, request, *args, **kwargs):
        group_name = request.POST.get('group_name')
        member_ids = request.POST.getlist('member_ids')
        remarks = request.POST.getlist('remarks') # This will be a list of remarks corresponding to member_ids

        if not group_name:
            messages.error(request, "Group name is required.")
            return self.get(request, *args, **kwargs)

        if not member_ids:
            messages.error(request, "Please select at least one member for the group.")
            return self.get(request, *args, **kwargs)

        # Create the new group
        group = ConstituentGroup.objects.create(
            name=group_name,
            description=f"Group created on {timezone.now().date()}", # Default description
            created_by=request.user,
            is_active=True
        )

        # Add members to the group and save remarks
        for i, member_id in enumerate(member_ids):
            member = get_object_or_404(BMParliamentMember, pk=member_id) # Fetch BMParliamentMember
            group.registrant_members.add(member) # Use the new ManyToManyField
            
            # Save remark for this member in this group (this requires a many-to-many through model)
            # For now, we'll just add a generic remark to the group description or log it.
            # A more robust solution would involve a custom intermediate model for ConstituentGroup.members
            # to store per-member remarks. For this task, we'll simplify.
            if remarks and i < len(remarks) and remarks[i]:
                # This is a simplified approach. In a real app, you'd likely have a dedicated model
                # to store remarks per member per group.
                logger.info(f"Remark for {member.get_full_name()} in group {group.name}: {remarks[i]}")

        messages.success(request, f"Group '{group.name}' created successfully with selected members.")
        return redirect('staff_constituent_group_detail', slug=group.slug)


class CheckMembershipView(TemplateView):
    """
    View for users to check their membership status.
    """
    template_name = 'constituents/check_membership.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('q', '').strip()
        context['search_query'] = search_query
        context['member_found'] = False
        context['member_data'] = None

        if search_query:
            # Normalize search query for better matching (e.g., remove extra spaces)
            normalized_query = ' '.join(search_query.split()).lower()

            # Annotate queryset with full name combinations for robust searching
            members = BMParliamentMember.objects.annotate(
                full_name_first_last=Concat('first_name', Value(' '), 'last_name'),
                full_name_last_first=Concat('last_name', Value(' '), 'first_name'),
                full_name_all=Concat('first_name', Value(' '), 'middle_name', Value(' '), 'last_name')
            ).filter(
                Q(first_name__icontains=normalized_query) |
                Q(last_name__icontains=normalized_query) |
                Q(middle_name__icontains=normalized_query) |
                Q(full_name_first_last__icontains=normalized_query) |
                Q(full_name_last_first__icontains=normalized_query) |
                Q(full_name_all__icontains=normalized_query)
            ).order_by('last_name', 'first_name').distinct()

            if members.exists():
                context['member_found'] = True
                context['members_list'] = []
                for member in members:
                    context['members_list'].append({
                        'full_name': member.get_full_name(),
                        'date_of_registration': member.date_of_application.strftime('%B %d, %Y') if member.date_of_application else 'N/A',
                        'status': member.get_status_display(),
                        'sector': member.get_sector_display() if member.sector else 'N/A',
                    })
                
                if members.count() > 1:
                    messages.info(self.request, f"Multiple members found for '{search_query}'. Please review the list below.")
            else:
                messages.info(self.request, "No membership found with that name. Please try again or contact our office.")
        
        return context


# Staff Constituent Views
class StaffConstituentDashboardView(StaffRequiredMixin, ListView):
    """
    Dashboard for staff to view and manage constituents.
    """
    model = Constituent
    template_name = 'constituents/staff/dashboard.html'
    context_object_name = 'constituents'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = Constituent.objects.all().select_related('user')
        
        # Filter by search query if provided
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) | 
                Q(user__last_name__icontains=search_query) | 
                Q(user__email__icontains=search_query) |
                Q(user__phone_number__icontains=search_query) |
                Q(voter_id__icontains=search_query)
            )
        
        # Filter by engagement level if provided
        engagement_level = self.request.GET.get('engagement_level')
        if engagement_level:
            if engagement_level == 'high':
                queryset = queryset.filter(engagement_level__gte=7)
            elif engagement_level == 'medium':
                queryset = queryset.filter(engagement_level__range=(4, 6))
            elif engagement_level == 'low':
                queryset = queryset.filter(engagement_level__range=(1, 3))
            elif engagement_level == 'inactive':
                queryset = queryset.filter(engagement_level=0)
        
        # Filter by voter status if provided
        is_voter = self.request.GET.get('is_voter')
        if is_voter:
            is_voter_bool = is_voter == 'true'
            queryset = queryset.filter(is_voter=is_voter_bool)
        
        # Filter by volunteer status if provided
        is_volunteer = self.request.GET.get('is_volunteer')
        if is_volunteer:
            is_volunteer_bool = is_volunteer == 'true'
            queryset = queryset.filter(is_volunteer=is_volunteer_bool)
        
        # Filter by gender if provided
        gender = self.request.GET.get('gender')
        if gender:
            queryset = queryset.filter(gender=gender)
        
        # Filter by group if provided
        group_id = self.request.GET.get('group')
        if group_id:
            queryset = queryset.filter(groups__id=group_id)
            
        # Sort results
        sort_by = self.request.GET.get('sort_by', '-created_at')
        if sort_by == 'name':
            queryset = queryset.order_by('user__first_name', 'user__last_name')
        elif sort_by == 'engagement_high':
            queryset = queryset.order_by('-engagement_level')
        elif sort_by == 'engagement_low':
            queryset = queryset.order_by('engagement_level')
        elif sort_by == 'recent_engagement':
            queryset = queryset.order_by('-last_engagement')
        else:
            queryset = queryset.order_by(sort_by)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get counts for sidebar filters
        context['engagement_counts'] = {
            'high': Constituent.objects.filter(engagement_level__gte=7).count(),
            'medium': Constituent.objects.filter(engagement_level__range=(4, 6)).count(),
            'low': Constituent.objects.filter(engagement_level__range=(1, 3)).count(),
            'inactive': Constituent.objects.filter(engagement_level=0).count(),
        }
        
        context['voter_counts'] = {
            'voters': Constituent.objects.filter(is_voter=True).count(),
            'non_voters': Constituent.objects.filter(is_voter=False).count(),
        }
        
        context['volunteer_counts'] = {
            'volunteers': Constituent.objects.filter(is_volunteer=True).count(),
            'non_volunteers': Constituent.objects.filter(is_volunteer=False).count(),
        }
        
        # Get all constituent groups for filtering
        context['groups'] = ConstituentGroup.objects.all().annotate(member_count=Count('members'))
        
        # Get selected filters for active state in UI
        context['selected_engagement_level'] = self.request.GET.get('engagement_level', '')
        context['selected_is_voter'] = self.request.GET.get('is_voter', '')
        context['selected_is_volunteer'] = self.request.GET.get('is_volunteer', '')
        context['selected_gender'] = self.request.GET.get('gender', '')
        context['selected_group'] = self.request.GET.get('group', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['sort_by'] = self.request.GET.get('sort_by', '-created_at')
        
        # Base params for pagination links
        filter_params = self.request.GET.copy()
        if 'page' in filter_params:
            del filter_params['page']
        context['filter_params'] = filter_params
        
        # Quick stats
        context['total_constituents'] = Constituent.objects.count()
        context['new_this_month'] = Constituent.objects.filter(
            created_at__gte=timezone.now().replace(day=1, hour=0, minute=0)
        ).count()
        context['avg_engagement'] = Constituent.objects.aggregate(avg=Avg('engagement_level'))['avg'] or 0
        
        return context
    
class StaffConstituentDetailView(StaffRequiredMixin, DetailView):
    """
    Detailed view of a constituent for staff members.
    """
    model = Constituent
    template_name = 'constituents/staff/detail.html'
    context_object_name = 'constituent'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        constituent = self.get_object()
        
        # Get all interactions for this constituent
        context['interactions'] = constituent.interactions.all().select_related('staff_member').order_by('-date')
        
        # Get all groups this constituent belongs to
        context['groups'] = constituent.groups.all()
        
        # Get referrals if available (assuming a relationship exists)
        try:
            from apps.referrals.models import Referral
            context['referrals'] = Referral.objects.filter(constituent=constituent.user).order_by('-created_at')
        except ImportError:
            # Referrals app might not be installed or no relationship exists
            context['referrals'] = []
        
        return context

class StaffConstituentCreateView(StaffRequiredMixin, CreateView):
    """
    Create a new constituent profile.
    """
    model = Constituent
    template_name = 'constituents/staff/form.html'
    fields = ['birth_date', 'gender', 'education_level', 'occupation', 'occupation_type', 
              'household_size', 'is_voter', 'voter_id', 'voting_center', 'alternate_contact', 
              'bio', 'profile_image', 'interests', 'language_preference', 'is_volunteer', 
              'volunteer_interests', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Constituent'
        context['submit_text'] = 'Create Constituent'
        return context
    
    def form_valid(self, form):
        # Get or create the user first
        email = self.request.POST.get('email')
        phone = self.request.POST.get('phone_number')
        first_name = self.request.POST.get('first_name')
        last_name = self.request.POST.get('last_name')
        
        # Check if user with this email already exists
        try:
            user = User.objects.get(email=email)
            # If user exists but already has a constituent profile, prevent duplicate
            if hasattr(user, 'constituent_profile'):
                messages.error(self.request, f"User with email {email} already has a constituent profile.")
                return self.form_invalid(form)
        except User.DoesNotExist:
            # Create new user
            user = User.objects.create_user(
                username=email,
                email=email,
                first_name=first_name,
                last_name=last_name,
                phone_number=phone,
                user_type='constituent'
            )
            
        # Now save the constituent profile
        constituent = form.save(commit=False)
        constituent.user = user
        constituent.created_at = timezone.now()
        constituent.updated_at = timezone.now()
        
        # Set membership date if they're becoming a member
        if self.request.POST.get('is_member') == 'on':
            constituent.membership_date = timezone.now().date()
            user.user_type = 'member'
            user.save()
        
        constituent.save()
        
        # Save to Notion if integration is available
        try:
            notion_service = NotionService()
            notion_id = notion_service.create_constituent(constituent)
            if notion_id:
                constituent.notion_id = notion_id
                constituent.save(update_fields=['notion_id'])
        except Exception as e:
            logger.error(f"Error saving constituent to Notion: {e}")
            
        messages.success(self.request, f"Constituent profile for {user.get_full_name()} created successfully.")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('staff_constituent_detail', kwargs={'pk': self.object.pk})

class StaffConstituentUpdateView(StaffRequiredMixin, UpdateView):
    """
    Update an existing constituent profile.
    """
    model = Constituent
    template_name = 'constituents/staff/form.html'
    fields = ['birth_date', 'gender', 'education_level', 'occupation', 'occupation_type', 
              'household_size', 'is_voter', 'voter_id', 'voting_center', 'alternate_contact', 
              'bio', 'profile_image', 'interests', 'language_preference', 'engagement_level',
              'is_volunteer', 'volunteer_interests', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Constituent'
        context['submit_text'] = 'Save Changes'
        # Pass user information to the template for display in the form
        context['user_data'] = self.object.user
        return context
    
    def form_valid(self, form):
        # Update user information if provided
        user = self.object.user
        if user:
            user.first_name = self.request.POST.get('first_name', user.first_name)
            user.last_name = self.request.POST.get('last_name', user.last_name)
            user.email = self.request.POST.get('email', user.email)
            user.phone_number = self.request.POST.get('phone_number', user.phone_number)
            
            # Update user type if membership status changes
            if self.request.POST.get('is_member') == 'on' and user.user_type == 'constituent':
                user.user_type = 'member'
                if not self.object.membership_date:
                    form.instance.membership_date = timezone.now().date()
            
            user.save()
        
        # Update the constituent record
        constituent = form.save(commit=False)
        constituent.updated_at = timezone.now()
        constituent.save()
        
        # Update in Notion if integration is available
        if constituent.notion_id:
            try:
                notion_service = NotionService()
                notion_service.update_constituent(constituent)
            except Exception as e:
                logger.error(f"Error updating constituent in Notion: {e}")
        
        messages.success(self.request, f"Constituent profile for {user.get_full_name()} updated successfully.")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('staff_constituent_detail', kwargs={'pk': self.object.pk})

class StaffConstituentInteractionCreateView(StaffRequiredMixin, CreateView):
    """
    Record a new interaction with a constituent.
    """
    model = ConstituentInteraction
    template_name = 'constituents/staff/interaction_form.html'
    fields = ['interaction_type', 'date', 'description', 'location', 'outcome', 'follow_up_date', 'follow_up_notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        constituent_id = self.kwargs.get('constituent_id')
        constituent = get_object_or_404(Constituent, id=constituent_id)
        context['constituent'] = constituent
        context['title'] = f'Record Interaction with {constituent.user.get_full_name()}'
        context['submit_text'] = 'Save Interaction'
        return context
    
    def form_valid(self, form):
        constituent_id = self.kwargs.get('constituent_id')
        constituent = get_object_or_404(Constituent, id=constituent_id)
        
        interaction = form.save(commit=False)
        interaction.constituent = constituent
        interaction.staff_member = self.request.user
        interaction.created_at = timezone.now()
        interaction.updated_at = timezone.now()
        interaction.save()
        
        # Update the constituent's last engagement date
        constituent.last_engagement = interaction.date.date()
        constituent.save(update_fields=['last_engagement'])
        
        # Update in Notion if integration is available
        try:
            notion_service = NotionService()
            notion_id = notion_service.create_constituent_interaction(interaction)
            if notion_id:
                interaction.notion_id = notion_id
                interaction.save(update_fields=['notion_id'])
        except Exception as e:
            logger.error(f"Error saving interaction to Notion: {e}")
        
        messages.success(self.request, f"Interaction with {constituent.user.get_full_name()} recorded successfully.")
        return redirect('staff_constituent_detail', pk=constituent.id)

class StaffConstituentGroupListView(StaffRequiredMixin, ListView):
    """
    List of constituent groups for staff management.
    """
    model = ConstituentGroup
    template_name = 'constituents/staff/group_list.html'
    context_object_name = 'groups'
    
    def get_queryset(self):
        return ConstituentGroup.objects.all().annotate(
            member_count=Count('members')
        ).order_by('-is_active', 'name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_groups_count'] = ConstituentGroup.objects.filter(is_active=True).count()
        context['total_groups_count'] = ConstituentGroup.objects.count()
        return context

class StaffConstituentGroupDetailView(StaffRequiredMixin, DetailView):
    """
    Detailed view of a constituent group.
    """
    model = ConstituentGroup
    template_name = 'constituents/staff/group_detail.html'
    context_object_name = 'group'
    slug_url_kwarg = 'slug'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.get_object()
        
        # Get all members of this group with their profiles
        context['members'] = group.registrant_members.all().select_related('user') # Changed to registrant_members
        
        # Calculate volunteer and voter counts
        context['volunteer_count'] = group.registrant_members.filter(is_volunteer=True).count()
        context['voter_count'] = group.registrant_members.filter(is_voter=True).count()
        
        return context

class StaffConstituentGroupCreateView(StaffRequiredMixin, CreateView):
    """
    Create a new constituent group.
    """
    model = ConstituentGroup
    template_name = 'constituents/staff/group_form.html'
    fields = ['name', 'description', 'is_active']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create New Group'
        context['submit_text'] = 'Create Group'
        return context
    
    def form_valid(self, form):
        group = form.save(commit=False)
        group.created_by = self.request.user
        group.created_at = timezone.now()
        group.updated_at = timezone.now()
        group.save()
        
        # Add initial members if provided
        member_ids = self.request.POST.getlist('members')
        if member_ids:
            constituents = Constituent.objects.filter(id__in=member_ids)
            group.members.add(*constituents)
        
        # Save to Notion if integration is available
        try:
            notion_service = NotionService()
            notion_id = notion_service.create_constituent_group(group)
            if notion_id:
                group.notion_id = notion_id
                group.save(update_fields=['notion_id'])
        except Exception as e:
            logger.error(f"Error saving group to Notion: {e}")
        
        messages.success(self.request, f"Constituent group '{group.name}' created successfully.")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('staff_constituent_group_detail', kwargs={'slug': self.object.slug})

class StaffConstituentGroupUpdateView(StaffRequiredMixin, UpdateView):
    """
    Update an existing constituent group.
    """
    model = ConstituentGroup
    template_name = 'constituents/staff/group_form.html'
    fields = ['name', 'description', 'is_active']
    slug_url_kwarg = 'slug'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Group'
        context['submit_text'] = 'Save Changes'
        
        # Get current members for display in the form
        context['current_members'] = self.object.members.all()
        return context
    
    def form_valid(self, form):
        group = form.save(commit=False)
        group.updated_at = timezone.now()
        group.save()
        
        # Update members if needed
        member_ids = self.request.POST.getlist('members')
        if member_ids:
            group.members.clear()
            constituents = Constituent.objects.filter(id__in=member_ids)
            group.members.add(*constituents)
        
        # Update in Notion if integration is available
        if group.notion_id:
            try:
                notion_service = NotionService()
                notion_service.update_constituent_group(group)
            except Exception as e:
                logger.error(f"Error updating group in Notion: {e}")
        
        messages.success(self.request, f"Constituent group '{group.name}' updated successfully.")
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('staff_constituent_group_detail', kwargs={'slug': self.object.slug})

class MemberCountReportView(DatabaseAccessMixin, TemplateView):
    """
    Comprehensive member count report showing general statistics and sector breakdowns.
    Supports both HTML and PDF formats.
    """
    template_name = 'constituents/member_count_report.html'

    def get(self, request, *args, **kwargs):
        # Check if PDF format is requested
        if request.GET.get('format') == 'pdf':
            return self.generate_pdf_report()
        return super().get(request, *args, **kwargs)

    def generate_pdf_report(self):
        # Get the same context data as HTML version
        context = self.get_context_data()

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        styles = getSampleStyleSheet()

        # Compact styles for single page
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#1f2937')
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=12,
            spaceAfter=8,
            textColor=colors.HexColor('#1f2937')
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=6,
            textColor=colors.HexColor('#374151')
        )

        # Build PDF content
        content = []

        # Title and timestamp
        content.append(Paragraph("Member Count Report", title_style))
        content.append(Paragraph(f"Generated on: {context['report_generated_at'].strftime('%B %d, %Y at %I:%M:%S %p')}", normal_style))
        content.append(Spacer(1, 10))

        # General Statistics - Compact
        content.append(Paragraph("General Statistics", heading_style))

        stats_data = [
            ['Metric', 'Count'],
            ['Total Members', str(context['total_members'])],
            ['Approved', str(context['approved_count'])],
            ['Pending', str(context['pending_count'])],
            ['Incomplete', str(context['incomplete_count'])],
            ['Non Compliant', str(context['non_compliant_count'])],
            ['This Month', str(context['this_month_count'])],
        ]

        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        content.append(stats_table)
        content.append(Spacer(1, 10))

        # Clean single column layout with proper sections
        if context['sector_stats']:
            content.append(Paragraph("SECTOR BREAKDOWN", heading_style))

            sector_data = [['Sector', 'Total', 'Approved', 'Pending', 'Incomplete', 'Non Compliant', '%']]
            for sector in context['sector_stats']:  # Show ALL sectors that have members
                sector_name_short = sector['name'][:30] + ('...' if len(sector['name']) > 30 else '')
                sector_data.append([
                    sector_name_short,
                    str(sector['total']),
                    str(sector['approved']),
                    str(sector['pending']),
                    str(sector['incomplete']),
                    str(sector['non_compliant']),
                    f"{sector['percentage']}%"
                ])

            sector_table = Table(sector_data)
            sector_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))

            content.append(sector_table)
            content.append(Spacer(1, 15))

        if context['province_stats']:
            content.append(Paragraph("PROVINCE BREAKDOWN", heading_style))

            province_data = [['Province', 'Total', 'Approved', 'Pending', 'Incomplete', 'Non Compliant']]
            for province in context['province_stats']:  # Show ALL provinces that have members
                province_data.append([
                    province['province'][:20] + ('...' if len(province['province']) > 20 else ''),
                    str(province['total']),
                    str(province['approved']),
                    str(province['pending']),
                    str(province['incomplete']),
                    str(province['non_compliant'])
                ])

            province_table = Table(province_data)
            province_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))

            content.append(province_table)
        else:
            content.append(Paragraph("No data available", normal_style))

        # Build PDF
        doc.build(content)

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="member_count_report_{context["report_generated_at"].strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get all members
        all_members = BMParliamentMember.objects.all()

        # General statistics
        context['total_members'] = all_members.count()
        context['approved_count'] = all_members.filter(status='approved').count()
        context['pending_count'] = all_members.filter(status='pending').count()
        context['incomplete_count'] = all_members.filter(status='incomplete').count()
        context['non_compliant_count'] = all_members.filter(status='non_compliant').count()
        context['denied_count'] = all_members.filter(status='denied').count()

        # This month's registrations
        context['this_month_count'] = all_members.filter(
            date_of_application__gte=timezone.now().replace(day=1, hour=0, minute=0)
        ).count()

        # Sector-wise breakdown
        sector_stats = []
        sector_choices = dict(BMParliamentMember.SECTOR_CHOICES)

        for sector_code, sector_name in BMParliamentMember.SECTOR_CHOICES:
            sector_members = all_members.filter(sector=sector_code)
            sector_total = sector_members.count()

            # Always include all sectors, even if they have no members yet
            sector_stats.append({
                'code': sector_code,
                'name': sector_name,
                'total': sector_total,
                'approved': sector_members.filter(status='approved').count(),
                'pending': sector_members.filter(status='pending').count(),
                'incomplete': sector_members.filter(status='incomplete').count(),
                'non_compliant': sector_members.filter(status='non_compliant').count(),
                'denied': sector_members.filter(status='denied').count(),
                'percentage': round((sector_total / context['total_members']) * 100, 2) if context['total_members'] > 0 else 0
            })

        # Sort by total count (descending)
        sector_stats.sort(key=lambda x: x['total'], reverse=True)
        context['sector_stats'] = sector_stats

        # Status distribution for pie chart data
        context['status_distribution'] = {
            'approved': context['approved_count'],
            'pending': context['pending_count'],
            'incomplete': context['incomplete_count'],
            'non_compliant': context['non_compliant_count'],
            'denied': context['denied_count']
        }

        # Location breakdown (province-wise) - Fixed to prevent duplicates
        province_stats_dict = {}

        # Get unique provinces and aggregate their data properly
        provinces = all_members.values('address_province').distinct()

        for province_data in provinces:
            province = province_data['address_province']
            if not province:  # Skip empty provinces
                continue

            province_members = all_members.filter(address_province=province)

            # Use province as key to avoid duplicates
            if province not in province_stats_dict:
                province_stats_dict[province] = {
                    'province': province,
                    'total': province_members.count(),
                    'approved': province_members.filter(status='approved').count(),
                    'pending': province_members.filter(status='pending').count(),
                    'incomplete': province_members.filter(status='incomplete').count(),
                    'non_compliant': province_members.filter(status='non_compliant').count(),
                }

        # Convert to list and sort by total count
        province_stats = list(province_stats_dict.values())
        province_stats.sort(key=lambda x: x['total'], reverse=True)
        context['province_stats'] = province_stats

        # Recent trends (last 6 months)
        six_months_ago = timezone.now() - timezone.timedelta(days=180)
        monthly_trends = all_members.filter(
            date_of_application__gte=six_months_ago
        ).extra(
            select={'month': "DATE_TRUNC('month', date_of_application)"}
        ).values('month').annotate(
            total=Count('id'),
            approved=Count('id', filter=Q(status='approved')),
            pending=Count('id', filter=Q(status='pending'))
        ).order_by('month')

        context['monthly_trends'] = list(monthly_trends)

        # Summary statistics
        context['report_generated_at'] = timezone.now()
        context['total_sectors_with_members'] = len([s for s in sector_stats if s['total'] > 0])

        return context


class StaffConstituentAnalyticsView(StaffRequiredMixin, TemplateView):
    """
    Analytics dashboard for constituent data.
    """
    template_name = 'constituents/staff/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get filter values from request
        province_filter = self.request.GET.get('province')
        municipality_filter = self.request.GET.get('municipality')
        sector_filter = self.request.GET.get('sector')

        # Base queryset
        queryset = BMParliamentMember.objects.filter(status='approved')

        # Apply filters
        if province_filter:
            queryset = queryset.filter(address_province=province_filter)
        if municipality_filter:
            queryset = queryset.filter(address_municipality=municipality_filter)
        if sector_filter:
            queryset = queryset.filter(sector=sector_filter)

        # Overall counts
        context['total_members'] = queryset.count()
        context['new_this_month'] = queryset.filter(
            approved_date__gte=timezone.now().replace(day=1)
        ).count()

        total_approved = BMParliamentMember.objects.filter(status='approved').count()
        if total_approved > 0:
            active_percentage = (queryset.count() / total_approved) * 100
        else:
            active_percentage = 0
        context['active_percentage'] = round(active_percentage, 2)

        # Registration trend (monthly)
        registration_trend = queryset.extra(select={'month': "DATE_TRUNC('month', date_of_application)"}).values('month').annotate(count=Count('id')).order_by('month')
        context['registration_trend_labels'] = json.dumps([entry['month'].strftime('%b %Y') for entry in registration_trend])
        context['registration_trend_data'] = json.dumps([entry['count'] for entry in registration_trend])

        # Location distribution
        location_distribution = queryset.values('address_province', 'address_municipality').annotate(count=Count('id')).order_by('-count')
        context['location_distribution'] = location_distribution

        # Sector distribution
        sector_distribution = queryset.values('sector').annotate(count=Count('id')).order_by('-count')
        sector_choices_dict = dict(BMParliamentMember.SECTOR_CHOICES)
        context['sector_distribution'] = [
            {'sector': sector_choices_dict.get(item['sector'], item['sector']), 'count': item['count']}
            for item in sector_distribution
        ]
        context['sector_distribution_labels'] = json.dumps([item['sector'] for item in context['sector_distribution']])
        context['sector_distribution_data'] = json.dumps([item['count'] for item in context['sector_distribution']])

        # Filter options
        context['provinces'] = BMParliamentMember.objects.values_list('address_province', flat=True).distinct()
        context['municipalities'] = BMParliamentMember.objects.values_list('address_municipality', flat=True).distinct()
        context['sectors'] = BMParliamentMember.SECTOR_CHOICES

        # Current filter values
        context['current_filters'] = {
            'province': province_filter,
            'municipality': municipality_filter,
            'sector': sector_filter,
        }

        return context

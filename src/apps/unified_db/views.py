from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView, View, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Count, Case, When, Value, CharField
from django.utils import timezone
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.core.exceptions import ValidationError
import logging
import json
import traceback
from difflib import SequenceMatcher

# Import models
from .models import GeneralDatabase, DatabaseEntry, DatabaseField, PersonLink
from apps.constituents.models import BMParliamentMember, Constituent

logger = logging.getLogger(__name__)


class HighAccessRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """
    Mixin that checks if the user has high-level access to use the unified database system.
    Authorized roles: Superuser, MP, Chief of Staff, System Admin, and Coordinator
    """

    def test_func(self):
        authorized_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
        return (self.request.user.is_superuser or
                self.request.user.user_type in authorized_roles)

    def handle_no_permission(self):
        """Handle permission denied cases appropriately."""
        if not self.request.user.is_authenticated:
            return super().handle_no_permission()
        else:
            messages.warning(self.request, 'You do not have permission to access the unified database system.')
            return redirect('home')


class UnifiedPersonSearchView(HighAccessRequiredMixin, TemplateView):
    """
    Main search interface for finding people across all databases.
    Prioritizes BMParliamentMember results as requested.
    """
    template_name = 'unified_db/person_search.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('q', '').strip()

        context['search_query'] = search_query
        context['search_results'] = []
        context['has_results'] = False
        context['result_count'] = 0

        if search_query:
            search_results = self.perform_unified_search(search_query)
            context['search_results'] = search_results
            context['has_results'] = len(search_results) > 0
            context['result_count'] = len(search_results)

            # Group results by person for better organization
            context['grouped_results'] = self.group_results_by_person(search_results)

        # Get available databases for filter options
        context['available_databases'] = GeneralDatabase.objects.filter(is_active=True)

        return context

    def perform_unified_search(self, query):
        """
        Perform unified search across all person-related models.
        Returns results in priority order: BMParliamentMember first, then others.
        """
        results = []

        # Normalize search query for better matching
        normalized_query = self.normalize_name(query)

        # Search BMParliamentMember first (highest priority)
        member_results = self.search_bm_parliament_members(query, normalized_query)
        results.extend(member_results)

        # Search Constituents
        constituent_results = self.search_constituents(query, normalized_query)
        results.extend(constituent_results)

        # Search Database Entries
        db_entry_results = self.search_database_entries(query, normalized_query)
        results.extend(db_entry_results)

        # Remove duplicates and sort by relevance
        unique_results = self.deduplicate_results(results)
        return sorted(unique_results, key=lambda x: x['relevance_score'], reverse=True)

    def search_bm_parliament_members(self, query, normalized_query):
        """Search BMParliamentMember records"""
        results = []

        # Build search query for BMParliamentMember
        search_filter = (
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(middle_name__icontains=query) |
            Q(email__icontains=query) |
            Q(contact_number__icontains=query) |
            Q(member_id__icontains=query)
        )

        members = BMParliamentMember.objects.select_related('user').filter(search_filter)

        for member in members:
            # Calculate relevance score
            relevance_score = self.calculate_relevance_score(
                query, normalized_query,
                member.first_name, member.last_name, member.middle_name,
                member.email, member.contact_number
            )

            # Debug logging to check if member data is correct
            full_name = member.get_full_name()
            logger.info(f"Processing BMParliamentMember: ID={member.id}, Name='{full_name}', First='{member.first_name}', Last='{member.last_name}', Middle='{member.middle_name}'")

            results.append({
                'type': 'bm_parliament_member',
                'id': member.id,
                'title': f"BM Parliament Member: {full_name}",
                'subtitle': f"Member ID: {member.member_id} | Sector: {member.get_sector_display()}",
                'description': f"Status: {member.get_status_display()} | Email: {member.email}",
                'url': reverse('database_registrant_detail', args=[member.id]),
                'relevance_score': relevance_score + 100,  # Boost BMParliamentMember results
                'data': member,
                'matched_field': 'BM Parliament Member Database'
            })

        return results

    def search_constituents(self, query, normalized_query):
        """Search Constituent records"""
        results = []

        # Build search query for Constituents
        search_filter = (
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(user__phone_number__icontains=query) |
            Q(voter_id__icontains=query)
        )

        constituents = Constituent.objects.select_related('user').filter(search_filter)

        for constituent in constituents:
            # Calculate relevance score
            relevance_score = self.calculate_relevance_score(
                query, normalized_query,
                constituent.user.first_name, constituent.user.last_name, '',
                constituent.user.email, getattr(constituent.user, 'phone_number', '')
            )

            # Get full name with fallback
            full_name = constituent.user.get_full_name()
            if not full_name.strip():
                full_name = constituent.user.username or f"User {constituent.user.id}"

            results.append({
                'type': 'constituent',
                'id': constituent.id,
                'title': f"Constituent: {full_name}",
                'subtitle': f"Voter ID: {constituent.voter_id or 'N/A'} | Engagement: {constituent.engagement_level}/10",
                'description': f"Email: {constituent.user.email} | Member since: {constituent.membership_date or 'N/A'}",
                'url': reverse('staff_constituent_detail', args=[constituent.id]),
                'relevance_score': relevance_score + 50,  # Boost constituent results but less than BM Parliament
                'data': constituent,
                'matched_field': 'Constituent Database'
            })

        return results

    def search_database_entries(self, query, normalized_query):
        """Search DatabaseEntry records"""
        results = []

        # Search in database entries - include entry_data search
        entries = DatabaseEntry.objects.select_related('database', 'linked_user').filter(
            Q(search_text__icontains=query.lower()) |
            Q(guest_first_name__icontains=query) |
            Q(guest_last_name__icontains=query) |
            Q(guest_email__icontains=query) |
            Q(entry_data__icontains=query)  # Search in JSON data
        )

        for entry in entries:
            # Calculate relevance score - also check entry_data for names
            first_name = entry.guest_first_name
            last_name = entry.guest_last_name
            email = entry.guest_email

            # If guest fields are empty, try to get names from entry_data
            if not first_name and not last_name:
                for key, value in entry.entry_data.items():
                    key_lower = key.lower()
                    if any(keyword in key_lower for keyword in ['first', 'fname', 'first_name']) and value:
                        first_name = str(value)
                    elif any(keyword in key_lower for keyword in ['last', 'lname', 'last_name', 'surname']) and value:
                        last_name = str(value)
                    elif any(keyword in key_lower for keyword in ['email', 'e-mail', 'mail']) and value:
                        email = str(value)

            relevance_score = self.calculate_relevance_score(
                query, normalized_query,
                first_name, last_name, entry.guest_middle_name,
                email, entry.guest_phone
            )

            # Get display name with better fallback logic
            display_name = entry.get_full_name()

            # If still no name, try to construct from available data
            if display_name in ["Guest User", "User"] and (first_name or last_name or email):
                name_parts = [p for p in [first_name, last_name] if p]
                if name_parts:
                    display_name = ' '.join(name_parts)
                elif email:
                    display_name = email.split('@')[0]  # Use email username as fallback
                else:
                    display_name = f"Entry #{entry.id}"

            # Generate safe URL for database entry - always use unified person detail for now
            try:
                # Use unified person detail view which handles all cases
                entry_url = reverse('unified_db:person_detail', args=['database_entry', entry.id])
            except:
                # Final fallback - just use a placeholder
                entry_url = "#"

            results.append({
                'type': 'database_entry',
                'id': entry.id,
                'title': f"Database Entry: {display_name}",
                'subtitle': f"Database: {entry.database.name} | Status: {entry.get_status_display()}",
                'description': f"Created: {entry.created_at.date()} | Contact: {email or 'N/A'}",
                'url': entry_url,
                'relevance_score': relevance_score,
                'data': entry,
                'matched_field': entry.database.name
            })

        return results

    def calculate_relevance_score(self, query, normalized_query, first_name, last_name, middle_name, email, phone):
        """Calculate relevance score based on name and field matching"""
        score = 0
        full_name = f"{first_name} {middle_name} {last_name}".strip()

        # Exact matches get highest scores
        if query.lower() == full_name.lower():
            score += 100
        elif query.lower() == f"{first_name} {last_name}".lower():
            score += 90
        elif query.lower() in full_name.lower():
            score += 70

        # Email matches
        if query.lower() in (email or '').lower():
            score += 60

        # Phone matches
        if query.lower() in (phone or '').lower():
            score += 50

        # Fuzzy matching using SequenceMatcher
        if full_name:
            name_similarity = SequenceMatcher(None, query.lower(), full_name.lower()).ratio()
            score += name_similarity * 40

        return score

    def normalize_name(self, name):
        """Normalize name for better matching"""
        import re
        # Remove extra spaces and convert to lowercase
        normalized = re.sub(r'\s+', ' ', name.strip().lower())
        return normalized

    def deduplicate_results(self, results):
        """Remove duplicate results based on person matching"""
        seen = set()
        unique_results = []

        for result in results:
            # Create a unique identifier for each person
            if result['type'] == 'bm_parliament_member':
                person_id = f"parliament_{result['data'].id}"
            elif result['type'] == 'constituent':
                person_id = f"constituent_{result['data'].id}"
            else:
                person_id = f"entry_{result['data'].id}"

            if person_id not in seen:
                seen.add(person_id)
                unique_results.append(result)

        return unique_results

    def group_results_by_person(self, results):
        """Group results by person for better display"""
        # This would implement more sophisticated person matching
        # For now, return results as-is with basic grouping
        return {
            'bm_parliament_members': [r for r in results if r['type'] == 'bm_parliament_member'],
            'constituents': [r for r in results if r['type'] == 'constituent'],
            'database_entries': [r for r in results if r['type'] == 'database_entry'],
        }


class UnifiedPersonDetailView(HighAccessRequiredMixin, DetailView):
    """
    Unified view showing all database entries for a specific person.
    Displays information from all linked databases.
    """
    template_name = 'unified_db/person_detail.html'
    context_object_name = 'person'

    def get_object(self, queryset=None):
        """Get person based on type and ID from URL parameters"""
        person_type = self.kwargs.get('person_type')
        person_id = self.kwargs.get('person_id')

        if person_type == 'bm_parliament_member':
            return BMParliamentMember.objects.select_related('user').get(id=person_id)
        elif person_type == 'constituent':
            return Constituent.objects.select_related('user').get(id=person_id)
        elif person_type == 'database_entry':
            return DatabaseEntry.objects.select_related('database', 'linked_user').get(id=person_id)
        else:
            raise ValueError(f"Unknown person type: {person_type}")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        person_obj = self.get_object()

        # Get all linked records for this person
        context['linked_records'] = self.get_linked_records(person_obj)
        context['person_links'] = self.get_person_links(person_obj)

        # Determine primary person info
        context['primary_info'] = self.get_primary_person_info(person_obj)

        return context

    def get_linked_records(self, person_obj):
        """Get all records linked to this person"""
        records = []

        if isinstance(person_obj, BMParliamentMember):
            records.append({
                'type': 'bm_parliament_member',
                'title': 'BM Parliament Member',
                'data': person_obj,
                'url': reverse('database_registrant_detail', args=[person_obj.id]),
                'is_primary': True
            })

            # Check for linked constituent
            if hasattr(person_obj, 'person_links'):
                for link in person_obj.person_links.filter(constituent__isnull=False):
                    records.append({
                        'type': 'constituent',
                        'title': 'Constituent Profile',
                        'data': link.constituent,
                        'url': reverse('staff_constituent_detail', args=[link.constituent.id]),
                        'is_primary': False
                    })

        elif isinstance(person_obj, Constituent):
            records.append({
                'type': 'constituent',
                'title': 'Constituent Profile',
                'data': person_obj,
                'url': reverse('staff_constituent_detail', args=[person_obj.id]),
                'is_primary': True
            })

        elif isinstance(person_obj, DatabaseEntry):
            records.append({
                'type': 'database_entry',
                'title': f'Database Entry - {person_obj.database.name}',
                'data': person_obj,
                'url': f"/databases/{person_obj.database.slug}/entries/{person_obj.id}/",
                'is_primary': True
            })

        return records

    def get_person_links(self, person_obj):
        """Get PersonLink objects for this person"""
        if isinstance(person_obj, BMParliamentMember):
            return PersonLink.objects.filter(bm_parliament_member=person_obj)
        elif isinstance(person_obj, Constituent):
            return PersonLink.objects.filter(constituent=person_obj)
        else:
            return PersonLink.objects.filter(database_entries=person_obj)

    def get_primary_person_info(self, person_obj):
        """Get primary information for display"""
        if isinstance(person_obj, BMParliamentMember):
            return {
                'name': person_obj.get_full_name(),
                'email': person_obj.email,
                'phone': person_obj.contact_number,
                'primary_type': 'BM Parliament Member',
                'member_id': person_obj.member_id,
                'sector': person_obj.get_sector_display(),
                'status': person_obj.get_status_display()
            }
        elif isinstance(person_obj, Constituent):
            return {
                'name': person_obj.user.get_full_name(),
                'email': person_obj.user.email,
                'phone': getattr(person_obj.user, 'phone_number', ''),
                'primary_type': 'Constituent',
                'voter_id': person_obj.voter_id,
                'engagement_level': person_obj.engagement_level
            }
        else:  # DatabaseEntry
            return {
                'name': person_obj.get_full_name(),
                'email': person_obj.guest_email,
                'phone': person_obj.guest_phone,
                'primary_type': f'Database Entry - {person_obj.database.name}',
                'database': person_obj.database.name,
                'status': person_obj.get_status_display()
            }


class DatabaseListView(HighAccessRequiredMixin, ListView):
    """
    List all available databases with management options.
    Shows hierarchical structure with parent-child relationships.
    """
    model = GeneralDatabase
    template_name = 'unified_db/database_list.html'
    context_object_name = 'databases'
    paginate_by = 20

    def get_queryset(self):
        queryset = GeneralDatabase.objects.filter(is_active=True).select_related()

        # Filter by type if specified
        db_type = self.request.GET.get('type')
        if db_type:
            queryset = queryset.filter(database_type=db_type)

        # Filter by search query
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        return queryset.annotate(
            entry_count=Count('entries'),
            sub_db_count=Count('sub_databases')
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get root databases (no parent) and build hierarchy
        root_databases = self.get_queryset().filter(parent_database__isnull=True)
        context['root_databases'] = root_databases
        context['database_hierarchy'] = self.build_database_hierarchy()

        # Statistics
        context['total_databases'] = GeneralDatabase.objects.filter(is_active=True).count()
        context['total_entries'] = DatabaseEntry.objects.count()
        context['root_database_count'] = root_databases.count()

        # Database type counts
        type_counts = GeneralDatabase.objects.filter(is_active=True).values('database_type').annotate(
            count=Count('id')
        )
        context['type_counts'] = {item['database_type']: item['count'] for item in type_counts}

        # Filter options
        context['current_filters'] = {
            'type': self.request.GET.get('type', ''),
            'search': self.request.GET.get('search', ''),
        }

        return context

    def build_database_hierarchy(self):
        """Build hierarchical structure for template display"""
        hierarchy = {}

        # Get all active databases with their relationships
        databases = GeneralDatabase.objects.filter(is_active=True).select_related('parent_database').annotate(
            entry_count=Count('entries'),
            sub_db_count=Count('sub_databases')
        )

        for db in databases:
            db_data = {
                'id': db.id,
                'name': db.name,
                'slug': db.slug,
                'description': db.description,
                'database_type': db.database_type,
                'entry_count': db.entry_count,
                'sub_db_count': db.sub_db_count,
                'created_at': db.created_at,
                'children': []
            }

            if db.parent_database:
                # This is a subdatabase - add to parent's children
                parent_slug = db.parent_database.slug
                if parent_slug not in hierarchy:
                    hierarchy[parent_slug] = {'children': []}
                hierarchy[parent_slug]['children'].append(db_data)
            else:
                # This is a root database
                hierarchy[db.slug] = db_data

        return hierarchy


class DatabaseCreateView(HighAccessRequiredMixin, CreateView):
    """
    Create a new custom database.
    """
    model = GeneralDatabase
    template_name = 'unified_db/database_form.html'
    fields = ['name', 'description', 'database_type', 'parent_database', 'is_public', 'allow_self_registration']

    def get_initial(self):
        initial = super().get_initial()
        # Pre-populate parent database if specified in URL
        parent_slug = self.request.GET.get('parent')
        if parent_slug:
            try:
                parent_db = GeneralDatabase.objects.get(slug=parent_slug, is_active=True)
                initial['parent_database'] = parent_db
            except GeneralDatabase.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create New Database'
        context['submit_text'] = 'Create Database'
        context['available_parents'] = GeneralDatabase.objects.filter(is_active=True)

        # Show context about parent database if specified
        parent_slug = self.request.GET.get('parent')
        if parent_slug:
            try:
                parent_db = GeneralDatabase.objects.get(slug=parent_slug, is_active=True)
                context['parent_database'] = parent_db
            except GeneralDatabase.DoesNotExist:
                pass

        return context

    def form_valid(self, form):
        database = form.save(commit=False)
        database.created_by = self.request.user
        database.save()

        # Set default access roles for high-level users
        if not database.allowed_roles:
            database.allowed_roles = ['superuser', 'mp', 'chief_of_staff', 'admin', 'coordinator']
            database.save()

        messages.success(self.request, f"Database '{database.name}' created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('unified_db:database_detail', args=[self.object.slug])


class DatabaseDetailView(HighAccessRequiredMixin, DetailView):
    """
    Detailed view of a database with entries and management options.
    """
    model = GeneralDatabase
    template_name = 'unified_db/database_detail.html'
    context_object_name = 'database'
    slug_url_kwarg = 'slug'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        database = self.get_object()

        # Get entries with pagination
        entries_list = database.entries.select_related('linked_user', 'created_by').all()
        paginator = Paginator(entries_list, 25)
        page_number = self.request.GET.get('page')
        entries = paginator.get_page(page_number)

        context['entries'] = entries
        context['fields'] = database.fields.order_by('order', 'name')

        # Statistics
        context['total_entries'] = database.entries.count()
        context['approved_entries'] = database.entries.filter(status='approved').count()
        context['pending_entries'] = database.entries.filter(status='draft').count()

        # Sub-databases
        context['sub_databases'] = database.sub_databases.filter(is_active=True)

        # Use active column names (only columns with data) instead of all column names
        context['active_columns'] = database.active_column_names

        return context


class DatabaseDeleteView(HighAccessRequiredMixin, DeleteView):
    """
    Delete a database with safety checks.
    Allows force deletion of databases with entries and sub-databases.
    """
    model = GeneralDatabase
    template_name = 'unified_db/database_confirm_delete.html'
    context_object_name = 'database'
    slug_url_kwarg = 'slug'

    def get_success_url(self):
        return reverse('unified_db:database_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        database = self.get_object()

        # Check for dependencies
        context['has_entries'] = database.entries.exists()
        context['has_sub_databases'] = database.sub_databases.exists()
        context['can_delete'] = True  # Allow deletion regardless of dependencies

        # Get counts for display
        context['entries_count'] = database.entries.count()
        context['sub_databases_count'] = database.sub_databases.count()

        # Calculate total items that will be deleted
        context['total_items'] = context['entries_count']
        if context['has_sub_databases']:
            # Count all entries in sub-databases recursively
            for sub_db in database.sub_databases.all():
                context['total_items'] += sub_db.entries.count()
                # Add sub-databases of sub-databases
                context['total_items'] += sub_db.sub_databases.count()

        return context

    def delete(self, request, *args, **kwargs):
        database = self.get_object()
        database_name = database.name

        # Get force delete parameter
        force_delete = request.POST.get('force_delete') == 'true'

        # Check if database has entries and not forcing deletion
        if database.entries.exists() and not force_delete:
            messages.warning(
                request,
                f"Database '{database.name}' contains {database.entries.count()} entries. "
                "Check the box to force delete and remove all entries as well."
            )
            return redirect('unified_db:database_delete', slug=database.slug)

        # Check if database has sub-databases and not forcing deletion
        if database.sub_databases.exists() and not force_delete:
            messages.warning(
                request,
                f"Database '{database.name}' has {database.sub_databases.count()} sub-databases. "
                "Check the box to force delete and remove all sub-databases as well."
            )
            return redirect('unified_db:database_delete', slug=database.slug)

        # Proceed with deletion (force delete if requested)
        try:
            if force_delete:
                # Delete all entries first
                entries_deleted = database.entries.count()
                database.entries.all().delete()

                # Delete all sub-databases recursively
                sub_databases_deleted = self.delete_sub_databases(database)

                # Delete the database itself
                database.delete()

                messages.success(
                    request,
                    f"Database '{database_name}' and all {entries_deleted} entries, "
                    f"{sub_databases_deleted} sub-databases have been deleted successfully."
                )
            else:
                # Simple deletion for empty databases
                database.delete()
                messages.success(request, f"Database '{database_name}' has been deleted successfully.")

        except Exception as e:
            messages.error(request, f"Error deleting database: {str(e)}")
            return redirect('unified_db:database_detail', slug=database.slug)

        return redirect('unified_db:database_list')

    def delete_sub_databases(self, database):
        """Recursively delete all sub-databases"""
        deleted_count = 0

        for sub_db in database.sub_databases.all():
            # Delete entries in sub-database
            sub_db.entries.all().delete()

            # Recursively delete sub-sub-databases
            deleted_count += self.delete_sub_databases(sub_db)

            # Delete the sub-database itself
            sub_db.delete()
            deleted_count += 1

        return deleted_count


class DatabaseEntryDetailView(HighAccessRequiredMixin, DetailView):
    """
    View details of a database entry.
    """
    model = DatabaseEntry
    template_name = 'unified_db/entry_detail.html'
    context_object_name = 'entry'

    def get_object(self, queryset=None):
        """Get the database entry"""
        database_slug = self.kwargs.get('database_slug')
        entry_id = self.kwargs.get('pk')

        # Try to find database by slug first, then by str if slug fails
        try:
            database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)
        except:
            # If slug lookup fails, try string lookup
            database = get_object_or_404(GeneralDatabase, name=database_slug, is_active=True)

        entry = get_object_or_404(DatabaseEntry, id=entry_id, database=database)

        return entry

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        entry = self.get_object()
        database = entry.database

        context['database'] = database
        context['fields'] = database.fields.order_by('order', 'name')
        context['title'] = f'Entry Details - {database.name}'

        return context


class DatabaseEntryCreateView(HighAccessRequiredMixin, CreateView):
    """
    Create a new entry in a database.
    """
    model = DatabaseEntry
    template_name = 'unified_db/entry_form.html'
    fields = ['linked_user', 'guest_first_name', 'guest_last_name', 'guest_middle_name', 'guest_email', 'guest_phone']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        database_slug = self.kwargs.get('database_slug')
        database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)

        context['database'] = database
        context['fields'] = database.fields.order_by('order', 'name')
        context['title'] = f'Add Entry to {database.name}'
        context['submit_text'] = 'Create Entry'

        return context

    def form_valid(self, form):
        database_slug = self.kwargs.get('database_slug')
        database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)

        entry = form.save(commit=False)
        entry.database = database
        entry.created_by = self.request.user

        # Validate custom fields
        custom_data = {}
        for field in database.fields.all():
            field_value = self.request.POST.get(f'field_{field.name}')
            try:
                field.validate_value(field_value)
                if field_value:
                    custom_data[field.name] = field_value
            except ValidationError as e:
                messages.error(self.request, f"Field '{field.label}': {e.message}")
                return self.form_invalid(form)

        entry.entry_data = custom_data
        entry.save()

        messages.success(self.request, f"Entry created successfully in {database.name}.")
        return super().form_valid(form)

    def get_success_url(self):
        database_slug = self.kwargs.get('database_slug')
        return reverse('unified_db:database_detail', args=[database_slug])


class DatabaseImportView(HighAccessRequiredMixin, View):
    """
    Handle Excel/CSV file import for bulk entry creation.
    """
    template_name = 'unified_db/import_preview.html'

    def get(self, request, database_slug):
        """Show import form"""
        database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)
        return render(request, self.template_name, {
            'database': database,
            'title': f'Import Data to {database.name}'
        })

    def post(self, request, database_slug):
        """Process uploaded file and show preview"""
        database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)

        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file to import.')
            return redirect('unified_db:database_import', database_slug=database_slug)

        excel_file = request.FILES['excel_file']

        # Validate file extension
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_extension = os.path.splitext(excel_file.name)[1].lower()

        if file_extension not in allowed_extensions:
            messages.error(request, 'Please upload a valid Excel file (.xlsx, .xls) or CSV file.')
            return redirect('unified_db:database_import', database_slug=database_slug)

        # Check if Excel packages are available for Excel files
        excel_packages_available = False
        if file_extension in ['.xlsx', '.xls']:
            try:
                import pandas
                excel_packages_available = True
            except ImportError:
                try:
                    import openpyxl
                    import xlrd
                    excel_packages_available = True
                except ImportError:
                    excel_packages_available = False

            if not excel_packages_available:
                messages.error(request, 'Excel file processing requires additional packages. Please use CSV format instead, or contact your administrator to install Excel processing packages (pandas/openpyxl/xlrd).')
                return redirect('unified_db:database_import', database_slug=database_slug)

        try:
            # Save file temporarily
            fs = FileSystemStorage()
            filename = fs.save(f'temp_import_{database.slug}_{request.user.id}{file_extension}', excel_file)
            file_path = fs.path(filename)

            # Read file based on type
            if file_extension == '.csv':
                # Read CSV file
                data = []
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        data.append(row)

                columns = list(data[0].keys()) if data else []
                preview_data = data[:10]  # First 10 rows for preview
                total_rows = len(data)

            else:
                # Read Excel file (.xlsx or .xls)
                try:
                    import pandas as pd

                    # Read all sheets from Excel file
                    excel_data = pd.read_excel(file_path, sheet_name=None)

                    # Get available sheet names
                    available_sheets = list(excel_data.keys())

                    # Validate available sheets
                    if not available_sheets:
                        messages.error(request, 'The Excel file contains no sheets.')
                        return redirect('unified_db:database_import', database_slug=database_slug)

                    # For now, use the first sheet as default, but we'll modify this to allow user selection
                    first_sheet_name = available_sheets[0]
                    df = excel_data[first_sheet_name]

                    # Convert to list of dictionaries
                    data = df.to_dict('records')

                    # Clean column names (remove extra spaces, handle NaN values)
                    if data:
                        # Get columns from the first row
                        columns = list(data[0].keys()) if data else []

                        # Clean data - replace NaN values with empty strings
                        for row in data:
                            for key, value in row.items():
                                if pd.isna(value):
                                    row[key] = ''
                                else:
                                    row[key] = str(value).strip()

                    preview_data = data[:10]  # First 10 rows for preview
                    total_rows = len(data)

                except ImportError:
                    # Try alternative Excel processing without pandas
                    try:
                        import xlrd
                        import openpyxl
                        # Use xlrd for .xls files and openpyxl for .xlsx files
                        if file_extension == '.xlsx':
                            wb = openpyxl.load_workbook(file_path, read_only=True)
                            ws = wb.active
                        else:  # .xls
                            wb = xlrd.open_workbook(file_path)
                            ws = wb.sheet_by_index(0)

                        data = []
                        headers = []

                        # Get headers from first row
                        if file_extension == '.xlsx':
                            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                        else:  # .xls
                            headers = [ws.cell_value(0, col) for col in range(ws.ncols)]

                        # Get data from subsequent rows
                        if file_extension == '.xlsx':
                            for row in ws.iter_rows(min_row=2):
                                row_data = {}
                                for i, cell in enumerate(row):
                                    if i < len(headers):
                                        row_data[headers[i]] = str(cell.value) if cell.value is not None else ''
                                data.append(row_data)
                        else:  # .xls
                            for row_idx in range(1, ws.nrows):
                                row_data = {}
                                for col_idx in range(ws.ncols):
                                    if row_idx < len(headers):
                                        row_data[headers[col_idx]] = str(ws.cell_value(row_idx, col_idx))
                                data.append(row_data)

                        # Clean data - replace None values with empty strings
                        for row in data:
                            for key, value in row.items():
                                if value is None:
                                    row[key] = ''

                        columns = headers
                        preview_data = data[:10]  # First 10 rows for preview
                        total_rows = len(data)

                    except ImportError:
                        messages.error(request, 'Excel processing requires additional packages. Please run: pip install pandas openpyxl xlrd')
                        return redirect('unified_db:database_import', database_slug=database_slug)
                    except Exception as e:
                        messages.error(request, f'Error reading Excel file: {str(e)}. Please ensure the file is not corrupted and try again.')
                        return redirect('unified_db:database_import', database_slug=database_slug)
                except Exception as e:
                    messages.error(request, f'Error reading Excel file: {str(e)}')
                    return redirect('unified_db:database_import', database_slug=database_slug)

            # Suggest field mappings based on column names
            field_suggestions = self.suggest_field_mappings(columns, database)

            return render(request, 'unified_db/import_preview.html', {
                'database': database,
                'columns': columns,
                'preview_data': preview_data,
                'total_rows': total_rows,
                'field_suggestions': field_suggestions,
                'existing_fields': database.fields.all(),
                'file_path': file_path,  # Pass file path to template
                'available_sheets': available_sheets if file_extension != '.csv' else None,
                'selected_sheet': first_sheet_name if file_extension != '.csv' else None,
                'title': f'Import Preview - {database.name}'
            })

        except Exception as e:
            # Clean up file if there's an error
            if 'file_path' in locals():
                try:
                    os.remove(file_path)
                except:
                    pass
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('unified_db:database_import', database_slug=database_slug)

    def suggest_field_mappings(self, columns, database):
        """Suggest automatic field mappings based on column names"""
        suggestions = {}

        # Common name patterns
        name_patterns = {
            'first_name': ['first name', 'firstname', 'first', 'fname', 'given name'],
            'last_name': ['last name', 'lastname', 'last', 'lname', 'surname', 'family name'],
            'middle_name': ['middle name', 'middlename', 'middle', 'mname'],
            'email': ['email', 'e-mail', 'email address', 'mail'],
            'phone': ['phone', 'phone number', 'mobile', 'contact', 'telephone'],
            'full_name': ['name', 'full name', 'complete name', 'last name, first name', 'lastname, firstname'],
        }

        # Check each column against patterns
        for col in columns:
            col_lower = col.lower().strip()
            suggestions[col] = {
                'suggested_field': 'custom',
                'confidence': 0,
                'field_name': col
            }

            # Check for full name patterns first (more specific)
            if any(pattern in col_lower for pattern in name_patterns['full_name']):
                suggestions[col] = {
                    'suggested_field': 'full_name',
                    'confidence': 95,
                    'field_name': 'full_name',
                    'parse_format': 'last_name_first_name'
                }
            # Check for other name fields
            else:
                for field_name, patterns in name_patterns.items():
                    if field_name != 'full_name':  # Skip full_name as we handled it above
                        for pattern in patterns:
                            if pattern in col_lower:
                                suggestions[col] = {
                                    'suggested_field': field_name,
                                    'confidence': 90,
                                    'field_name': field_name
                                }
                                break

            # Check for existing database fields
            for field in database.fields.all():
                if field.name.lower() in col_lower or col_lower in field.name.lower():
                    suggestions[col] = {
                        'suggested_field': 'existing',
                        'confidence': 95,
                        'field_name': field.name,
                        'field_id': field.id
                    }
                    break

        return suggestions


class DatabaseImportProcessView(HighAccessRequiredMixin, View):
    """
    Process the Excel import after user confirmation.
    """

    def post(self, request, database_slug):
        """Process the import with user-specified field mappings"""
        database = get_object_or_404(GeneralDatabase, slug=database_slug, is_active=True)

        # Get import configuration
        file_path = request.POST.get('file_path')
        selected_sheet = request.POST.get('selected_sheet')
        field_mappings_json = request.POST.get('field_mappings_json', '{}')

        logger.info(f"Received file_path: {file_path}")
        logger.info(f"Received field_mappings_json: {field_mappings_json}")

        # Parse field mappings
        try:
            field_mappings = json.loads(field_mappings_json)
            logger.info(f"Parsed field mappings: {field_mappings}")
        except json.JSONDecodeError as e:
            logger.warning(f"JSON decode error: {e}")
            # Try alternative parsing
            field_mappings_str = request.POST.get('field_mappings', '{}')
            try:
                field_mappings = json.loads(field_mappings_str)
            except json.JSONDecodeError:
                field_mappings = {}

        import_mode = request.POST.get('import_mode', 'preview')

        if not file_path or not os.path.exists(file_path):
            messages.error(request, 'Import file not found. Please try again.')
            return redirect('unified_db:database_import', database_slug=database_slug)

        try:
            # Read file based on extension
            file_extension = os.path.splitext(file_path)[1].lower()

            if file_extension == '.csv':
                # Read CSV file
                data = []
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        data.append(row)
            else:
                # Read Excel file
                try:
                    import pandas as pd

                    # Read all sheets to validate selected_sheet
                    excel_data = pd.read_excel(file_path, sheet_name=None)
                    available_sheets = list(excel_data.keys())

                    # Validate selected sheet
                    if selected_sheet and selected_sheet not in available_sheets:
                        messages.error(request, f'Selected sheet "{selected_sheet}" not found in the Excel file.')
                        return redirect('unified_db:database_import', database_slug=database_slug)

                    # Use selected sheet or default to first sheet
                    sheet_to_use = selected_sheet if selected_sheet else available_sheets[0]
                    df = excel_data[sheet_to_use]

                    data = df.to_dict('records')

                    # Clean data - replace NaN values with empty strings
                    for row in data:
                        for key, value in row.items():
                            if pd.isna(value):
                                row[key] = ''
                            else:
                                row[key] = str(value).strip()

                except ImportError:
                    # Try alternative Excel processing without pandas
                    try:
                        import xlrd
                        import openpyxl

                        # Use xlrd for .xls files and openpyxl for .xlsx files
                        if file_extension == '.xlsx':
                            wb = openpyxl.load_workbook(file_path, read_only=True)
                            ws = wb.active
                        else:  # .xls
                            wb = xlrd.open_workbook(file_path)
                            ws = wb.sheet_by_index(0)

                        data = []
                        headers = []

                        # Get headers from first row
                        if file_extension == '.xlsx':
                            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                        else:  # .xls
                            headers = [ws.cell_value(0, col) for col in range(ws.ncols)]

                        # Get data from subsequent rows
                        if file_extension == '.xlsx':
                            for row in ws.iter_rows(min_row=2):
                                row_data = {}
                                for i, cell in enumerate(row):
                                    if i < len(headers):
                                        row_data[headers[i]] = str(cell.value) if cell.value is not None else ''
                                data.append(row_data)
                        else:  # .xls
                            for row_idx in range(1, ws.nrows):
                                row_data = {}
                                for col_idx in range(ws.ncols):
                                    if row_idx < len(headers):
                                        row_data[headers[col_idx]] = str(ws.cell_value(row_idx, col_idx))
                                data.append(row_data)

                        # Clean data - replace None values with empty strings
                        for row in data:
                            for key, value in row.items():
                                if value is None:
                                    row[key] = ''

                    except ImportError:
                        messages.error(request, 'Excel processing requires additional packages. Please run: pip install pandas openpyxl xlrd')
                        return redirect('unified_db:database_import', database_slug=database_slug)
                    except Exception as e:
                        messages.error(request, f'Error reading Excel file: {str(e)}. Please ensure the file is not corrupted and try again.')
                        return redirect('unified_db:database_import', database_slug=database_slug)
                except Exception as e:
                    messages.error(request, f'Error reading Excel file: {str(e)}')
                    return redirect('unified_db:database_import', database_slug=database_slug)

            # Process the import
            import_results = self.process_import(data, database, field_mappings, request.user)

            # Clean up file
            os.remove(file_path)

            # Show results
            return render(request, 'unified_db/import_results.html', {
                'database': database,
                'import_results': import_results,
                'title': f'Import Results - {database.name}'
            })

        except Exception as e:
            messages.error(request, f'Error during import: {str(e)}')
            return redirect('unified_db:database_import', database_slug=database_slug)

    def process_import(self, data, database, field_mappings, user):
        """Process the actual import of data"""
        results = {
            'total_rows': len(data),
            'successful_imports': 0,
            'failed_imports': 0,
            'errors': [],
            'imported_entries': []
        }

        # Get existing fields for mapping
        existing_fields = {field.name: field for field in database.fields.all()}

        for index, row in enumerate(data):
            try:
                # Create entry data from row
                entry_data = {}

                # Map standard fields
                guest_first_name = ''
                guest_last_name = ''
                guest_middle_name = ''
                guest_email = ''
                guest_phone = ''

                # Debug logging
                logger.info(f"Processing row {index + 1}: {row}")
                logger.info(f"Field mappings: {field_mappings}")

                for excel_col, field_config in field_mappings.items():
                    logger.info(f"Processing column '{excel_col}' with config {field_config}")

                    # Check if column exists in row data
                    if excel_col in row:
                        raw_value = row[excel_col]
                        logger.info(f"Raw value for {excel_col}: '{raw_value}'")

                        # Check if value is not None and not empty
                        if raw_value is not None and str(raw_value).strip():
                            value = str(raw_value).strip()
                            logger.info(f"Processed value: '{value}'")

                            try:
                                # Parse field_config if it's a string
                                if isinstance(field_config, str):
                                    field_config = json.loads(field_config)

                                logger.info(f"Parsed field config: {field_config}")

                                # Handle field mapping based on type
                                field_type = field_config.get('type')
                                field_name = field_config.get('field')

                                logger.info(f"Field type: {field_type}, Field name: {field_name}")

                                if field_type == 'standard':
                                    logger.info(f"Mapping to standard field: {field_name}")

                                    if field_name == 'first_name':
                                        guest_first_name = value
                                        logger.info(f"Set guest_first_name: {guest_first_name}")
                                    elif field_name == 'last_name':
                                        guest_last_name = value
                                        logger.info(f"Set guest_last_name: {guest_last_name}")
                                    elif field_name == 'middle_name':
                                        guest_middle_name = value
                                        logger.info(f"Set guest_middle_name: {guest_middle_name}")
                                    elif field_name == 'full_name':
                                        # Parse full name in "Last Name, First Name Middle Initial" format
                                        logger.info(f"Parsing full name: {value}")
                                        parsed_names = self.parse_full_name(value)
                                        guest_first_name = parsed_names['first_name']
                                        guest_last_name = parsed_names['last_name']
                                        guest_middle_name = parsed_names['middle_name']
                                        logger.info(f"Parsed - First: '{guest_first_name}', Last: '{guest_last_name}', Middle: '{guest_middle_name}'")
                                    elif field_name == 'email':
                                        guest_email = value
                                        logger.info(f"Set guest_email: {guest_email}")
                                    elif field_name == 'phone':
                                        guest_phone = value
                                        logger.info(f"Set guest_phone: {guest_phone}")

                                elif field_type == 'existing_field':
                                    # Map to existing database field
                                    if field_name in existing_fields:
                                        entry_data[field_name] = value
                                        logger.info(f"Set existing field {field_name}: {value}")
                                    else:
                                        logger.warning(f"Existing field {field_name} not found in database fields")

                                elif field_type == 'new_field':
                                    # Create new field data
                                    entry_data[field_name] = value
                                    logger.info(f"Set new field {field_name}: {value}")

                                else:
                                    logger.warning(f"Unknown field type: {field_type}")

                            except json.JSONDecodeError as e:
                                logger.warning(f"JSON decode error for {excel_col}: {e}")
                                # Try to handle as direct field mapping
                                col_lower = excel_col.lower()
                                if any(keyword in col_lower for keyword in ['first', 'fname', 'first_name']):
                                    guest_first_name = value
                                elif any(keyword in col_lower for keyword in ['last', 'lname', 'last_name', 'surname']):
                                    guest_last_name = value
                                elif any(keyword in col_lower for keyword in ['middle', 'mname', 'middle_name']):
                                    guest_middle_name = value
                                elif any(keyword in col_lower for keyword in ['email', 'e-mail', 'mail']):
                                    guest_email = value
                                elif any(keyword in col_lower for keyword in ['phone', 'phone_number', 'mobile', 'contact', 'telephone']):
                                    guest_phone = value
                                else:
                                    entry_data[excel_col] = value

                logger.info(f"Final values - First: '{guest_first_name}', Last: '{guest_last_name}', Email: '{guest_email}'")

                # Validate that we have at least some data
                has_data = any([
                    guest_first_name.strip(),
                    guest_last_name.strip(),
                    guest_email.strip(),
                    guest_phone.strip(),
                    entry_data
                ])

                if not has_data:
                    # Try fallback auto-detection if field mapping didn't work
                    logger.warning(f"Row {index + 1} field mapping failed, trying auto-detection. Row data: {row}")

                    # Auto-detect common field patterns
                    for col_name, value in row.items():
                        if value and str(value).strip():
                            col_lower = col_name.lower().strip()

                            # Auto-detect name fields
                            if any(keyword in col_lower for keyword in ['first', 'fname', 'first_name']):
                                if not guest_first_name.strip():
                                    guest_first_name = str(value).strip()
                            elif any(keyword in col_lower for keyword in ['last', 'lname', 'last_name', 'surname']):
                                if not guest_last_name.strip():
                                    guest_last_name = str(value).strip()
                            elif any(keyword in col_lower for keyword in ['middle', 'mname', 'middle_name']):
                                if not guest_middle_name.strip():
                                    guest_middle_name = str(value).strip()
                            elif any(keyword in col_lower for keyword in ['email', 'e-mail', 'mail']):
                                if not guest_email.strip():
                                    guest_email = str(value).strip()
                            elif any(keyword in col_lower for keyword in ['phone', 'mobile', 'contact', 'telephone']):
                                if not guest_phone.strip():
                                    guest_phone = str(value).strip()
                            else:
                                # Store other fields as custom data
                                entry_data[col_name] = str(value).strip()

                    # Check again after auto-detection
                    has_data = any([
                        guest_first_name.strip(),
                        guest_last_name.strip(),
                        guest_email.strip(),
                        guest_phone.strip(),
                        entry_data
                    ])

                if not has_data:
                    logger.warning(f"Row {index + 1} has no valid data even after auto-detection. Row data: {row}")
                    results['failed_imports'] += 1
                    results['errors'].append({
                        'row': index + 1,
                        'error': 'No valid data found in row'
                    })
                    continue

                # Create the database entry
                entry = DatabaseEntry.objects.create(
                    database=database,
                    created_by=user,
                    guest_first_name=guest_first_name,
                    guest_last_name=guest_last_name,
                    guest_middle_name=guest_middle_name,
                    guest_email=guest_email,
                    guest_phone=guest_phone,
                    entry_data=entry_data,
                    status='submitted'
                )

                logger.info(f"Created entry with name: {entry.get_full_name()}")

                results['successful_imports'] += 1
                results['imported_entries'].append({
                    'entry': entry,
                    'row_number': index + 1
                })

            except Exception as e:
                logger.error(f"Error processing row {index + 1}: {str(e)}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                results['failed_imports'] += 1
                results['errors'].append({
                    'row': index + 1,
                    'error': str(e)
                })

        return results

    def parse_full_name(self, full_name):
        """
        Parse full name in "Last Name, First Name Middle Initial" format
        Returns dict with first_name, last_name, middle_name
        """
        first_name = ''
        last_name = ''
        middle_name = ''

        if not full_name or not full_name.strip():
            return {
                'first_name': first_name,
                'last_name': last_name,
                'middle_name': middle_name
            }

        # Clean up the full name
        name = full_name.strip()

        # Handle "Last Name, First Name Middle Initial" format
        if ',' in name:
            parts = [p.strip() for p in name.split(',')]
            if len(parts) >= 2:
                last_name = parts[0]
                first_and_middle = parts[1]

                # Split first name and middle initial
                name_parts = first_and_middle.split()
                if len(name_parts) >= 1:
                    first_name = name_parts[0]
                if len(name_parts) >= 2:
                    middle_name = ' '.join(name_parts[1:])  # Could be multiple middle names
        else:
            # Handle "First Name Last Name" format or single names
            name_parts = name.split()
            if len(name_parts) == 1:
                # Single name - could be first or last name
                first_name = name_parts[0]
            elif len(name_parts) == 2:
                # First and last name
                first_name = name_parts[0]
                last_name = name_parts[1]
            elif len(name_parts) >= 3:
                # First, middle, and last name
                first_name = name_parts[0]
                middle_name = ' '.join(name_parts[1:-1])
                last_name = name_parts[-1]

        return {
            'first_name': first_name,
            'last_name': last_name,
            'middle_name': middle_name
        }


# Import Paginator for pagination
from django.core.paginator import Paginator
from django.core.files.storage import FileSystemStorage
import os
import csv
